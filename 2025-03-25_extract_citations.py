# This is extract_citations.py


import requests
import PyPDF2
import re
import os
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter




def download_pdf(url):
    """Downloads a PDF from the given URL and saves it as a temporary file."""
    try:
        response = requests.get(url, stream=True, timeout=10)
        response.raise_for_status()
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        for chunk in response.iter_content(chunk_size=1024):
            temp_file.write(chunk)
        temp_file.close()
        print(f"Downloaded {url}")
        return temp_file.name
    except Exception as e:
        print(f"Failed to download {url}: {e}")
        return None




def sanitize_text(text):
    """Removes line breaks and carriage returns from a given text."""
    return re.sub(r"[\r\n]+", " ", text).strip()




def clean_citation(citation):
    """Cleans up citation syntax to match standard formats."""
    citation = re.sub(r"\b(\d+)\s*(U\.S\.C\.|USC)\s*(\d+)\b", r"\1 USC \3", citation)
    citation = re.sub(r"\b(\d+)\s*(C\.F\.R\.|CFR)\s*(\d+)\b", r"\1 CFR \3", citation)
    citation = re.sub(r"\b(E\.O\.|Executive\s*Order|Exec\.? Order)\s*(\d+)\b", r"Executive Order \2", citation)
    citation = re.sub(r"\bEO\s+(\d+)\b", r"Executive Order \1", citation)
    return citation




def extract_toc(reader):
    """Attempts to extract the Table of Contents (TOC) from the PDF."""
    toc = []
    toc_pattern = r"(?P<heading>.+?)\s+(\d+)"
    for page_num, page in enumerate(reader.pages[:10]):
        text = page.extract_text()
        if text and "Table of Contents" in text:
            matches = re.findall(toc_pattern, text)
            for match in matches:
                heading = sanitize_text(match[0])
                page_start = int(match[1])
                toc.append((heading, page_start))
    return toc




def infer_section_name(toc, page_num, context, page_text):
    """Infers the section name based on TOC or contextual headers."""
    if toc:
        for i, (section, start_page) in enumerate(toc):
            if i + 1 < len(toc) and toc[i + 1][1] > page_num >= start_page:
                return section
            elif i == len(toc) - 1 and page_num >= start_page:
                return section


    # Fallback: Find the nearest one-line paragraph before the context
    lines = page_text.splitlines()
    context_start = page_text.find(context)
    for i in range(len(lines) - 1, -1, -1):
        if len(lines[i].strip()) > 0 and lines[i].strip() in page_text[:context_start]:
            return sanitize_text(lines[i])
    return "Unknown Section"




def extract_us_code_citations(pdf_path, url):
    """Extracts U.S. Code, CFR, EO, OMB, Public Law, and Statutes citations from a PDF."""
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            toc = extract_toc(reader)
            num_pages = len(reader.pages)


            citations = []
            citation_pattern = r"""(?ix)
                (\b\d+\s*(U\.S\.C\.|USC|U\.S\. Code)\s*\u00a7?\s*\d+(\.\d+)*[a-zA-Z0-9]*\b) |
                (\b\d+\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*\u00a7?\s*\d+(\.\d+)*[a-zA-Z0-9]*\b) |
                (\b(E\.O\.|Executive\s*Order|EO|Exec\.? Order)\s*\d+\b) |
                (\bOMB\b|\bO\.M\.B\.\b|\bOMB Guidance\b|\bOffice of Management and Budget\b|\bOMB Circular\b|\bOMB Memo\b) |
                (\bPublic Law\b|\bP\.L\.\b|\bStatutes at Large\b|\bFederal Law\b) |
                (\bC\.?F\.?R\.?\b|\bCFR\b|\bC\.FR\b|\bCF\.R\b|\bCFR\.\b|\bC\.F\.R\b|\bC.F.R.\b) |
                (\bUSC\b|\bU\.S\.C\.\b|\bUS Code\b|\bU\.S\. Code\b|\bUnited States Code\b) |
                (\bCode of Federal Regulations\b|\bCFR Title\b|\bFederal Register\b) |
                (\b\d{1,5}\s+Stat\.\s+\d{1,5}\b)
            """


            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()
                if not text:
                    continue


                matches = re.finditer(citation_pattern, text, re.VERBOSE | re.IGNORECASE)


                for match in matches:
                    citation_text = match.group(0)
                    citation = clean_citation(citation_text)
                    start, end = match.start(), match.end()
                    context = sanitize_text(text[max(0, start - 100):min(len(text), end + 100)])
                    section_name = infer_section_name(toc, page_num + 1, context, text)
                    citation_page_url = f"{url}#page={page_num + 1}"
                    citations.append((citation, citation_page_url, section_name, context, url))


            return citations
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return []




def process_url(url):
    """Handles downloading, extracting, and cleaning up for a single URL."""
    temp_file = download_pdf(url)
    if not temp_file:
        return []


    try:
        citations = extract_us_code_citations(temp_file, url)
    finally:
        os.remove(temp_file)
    return citations




def save_to_excel(data, filename="extracted_citations.xlsx"):
    """Saves the extracted data to an Excel file."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Citation", "Citation Page", "Inferred Section Name", "Context", "URL"])


    for row in data:
        sanitized_row = [sanitize_text(str(cell)) for cell in row]
        sanitized_row[0] = clean_citation(sanitized_row[0])
        sheet.append(sanitized_row)


    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 20


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    workbook.save(filename)
    print(f"Saved data to {filename}")




def main():
    url_list = [
        "https://www.dhs.gov/sites/default/files/2024-09/2024_0923_cio_dhs_compliance_plan_omb_memoranda.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_047-01-privacy-policy-and-compliance_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/05.%20Directive%20138-01%2C%20Enterprise%20Information%20Technology%20Configuration%20Management%20%285-6-14%29.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_138-03-info-tech-asset-mgmt-and-refresh_revision-00.pdf",
        "https://www.dhs.gov/sites/default/files/publications/mgmt/information-and-technology-management/mgmt-dir_139-02-info-quality_revision-01.pdf"
        # Add more URLs as needed
    ]


    all_citations = []
    for url in url_list:
        all_citations.extend(process_url(url))


    save_to_excel(all_citations)




if __name__ == "__main__":
    main()
